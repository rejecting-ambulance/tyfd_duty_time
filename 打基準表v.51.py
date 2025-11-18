#網頁操作
from selenium import webdriver
from selenium.webdriver.support.ui import Select #選單
from selenium.webdriver.common.by import By #定位
from selenium.webdriver.support.ui import WebDriverWait #等待載入
from selenium.webdriver.support import expected_conditions as EC #等待載入
#excel操作
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
#輔助
import calendar
import time
import sys
import os
################版本資訊################
##################5.1##################
################函數大大們################
#補0
def add_zero(cc):
    if len(cc) < 2:
        cc = '0' + cc
    return cc

#警告處理
def alert_click():
    try:
        time.sleep(0.2)
        alert = driver.switch_to.alert
        print(f'  {alert.text}')
        alert.accept()
    except:
        None
    return 0

def get_exe_dir():
    """ 取得 `.exe` 真正所在的目錄 """
    if getattr(sys, 'frozen', False):  # PyInstaller 打包後
        return os.path.dirname(sys.executable)
    else:
        return os.path.dirname(os.path.abspath(__file__))

#系統檢測
def excel_check(file_name):
    if not os.path.exists(file_name):
        return "檔案不存在"

    try:
        with open(file_name, "a"):  # 嘗試以追加模式打開
            return "檔案可讀取"
    except PermissionError:
        return "檔案被占用"

# Excel 基本資料讀取
def data_loading(file_name, user_book):
    wb = load_workbook(file_name, data_only=True)
    ws = wb[user_book]

    user_account = str(ws['J2'].value)
    user_password = str(ws['J3'].value)
    user_name = str(ws['J4'].value)
    ex_year = str(ws['J5'].value - 1911)
    ex_month = add_zero(str(ws['J6'].value))
    day_number = calendar.monthrange(ws['J5'].value, ws['J6'].value)[1] 
    is_del = str(ws['J8'].value)

    if user_account == "None" or user_password == "None" or user_name == "None":
        print('基本資料不完全，請關閉於Excel或此輸入\n')
        time.sleep(0.5)
        user_account = input('請輸入帳號↵\n')
        user_password = input('請輸入密碼↵\n')
        user_name = input('請輸入姓名↵\n')

    return {
        "name": user_name,
        "account": user_account,
        "password": user_password,
        "month": ex_month,
        "year": ex_year,
        "days": day_number,
        "isdel": is_del
    }

# 各假別內容讀取(初始、自訂)
def setting_loading(file_name, user_book):
    wb = load_workbook(file_name, data_only=True)
    ws = wb[user_book]

    duty_dict = {}
    for i in range(2, 39, 6):  # 每 6 行一組
        duty_name = str(ws[f'B{i}'].value)
        if duty_name == 'None':  
            break

        rest_list = []
        for p in range(i, i + 6):
            temp = str(ws[f'C{p}'].value)
            if temp == 'None':
                break

            # 使用列表推導式收集數據
            rest_details = [
                str(ws[f'{get_column_letter(q+2)}{p}'].value) or ""  # 若為 None 則設為空字串
                for q in range(1, 5)
            ]
            rest_list.append(rest_details)

        duty_dict[duty_name] = rest_list
    
    for i in range(2, 39, 6):  # 每 6 行一組
        duty_name = str(ws[f'H{i}'].value)
        if duty_name == 'None':  
            break

        rest_list = []
        for p in range(i, i + 6):
            temp = str(ws[f'I{p}'].value)
            if temp == 'None':
                break

            # 使用列表推導式收集數據
            rest_details = [
                str(ws[f'{get_column_letter(q+8)}{p}'].value) or ""  # 若為 None 則設為空字串
                for q in range(1, 5)
            ]
            rest_list.append(rest_details)

        duty_dict[duty_name] = rest_list
    return duty_dict

#假別內容讀取
def duty_loading(file_name, user_book, days):
    wb = load_workbook(file_name, data_only=True)
    ws = wb[user_book]

    for i in range(days):
        d = i+2
        date = str(i+1)

        #將班別.A起A迄.B起B迄整理為陣列(C~G欄)
        arr = [
                str(ws[f'{get_column_letter(q+3)}{d}'].value) or ""  # 若為 None 則設為空字串
                for q in range(5)
            ]
        #arr[0]：班別
        if arr[0] == 'None':
            continue

        print(f'輸入 {i+1} 日： {arr[0]}')
        time_function(date, duty_dict[arr[0]], arr)

    return 0

#輸入函式(輸入內容陣列、日期)
def time_function(date_num,cont_arr, day_arr):
    #ID_arr = ['_selOFFTYPE', '_selLEAVETYPE', '_selTASKDATE', '_selDATES', '_selTASKHOURS', '_selDATEE', '_selTASKHOURE']
    for p in range(len(cont_arr)):
        day_or_hour = cont_arr[p][0]
        kind = cont_arr[p][1]
        date_num_st = 1
        date_num_en = 1
        time_st = cont_arr[p][2]
        time_en = cont_arr[p][3]

        #自訂時數帶入、隔日時數處理
        if time_st == 'A':
            time_st = day_arr[1] 
        elif time_st == 'B':
            time_st = day_arr[3]

        if time_en == 'A':
            time_en = day_arr[2]
        elif time_en == 'B':
            time_en = day_arr[4]
        
        try:    #用try是為了排除time_st == 'None'
            if int(time_st) < 8:
                date_num_st = 2
        except:
            date_num_st = 1
        try:
            if int(time_en) <= 8:
                date_num_en = 2
        except:
            date_num_en = 1

        #數字標準化
        date_num = add_zero(date_num)   
        time_st = add_zero(time_st)
        time_en = add_zero(time_en)

        #假別單位
        try:
            dropdown1 = wait.until(EC.visibility_of_element_located((By.ID, '_selOFFTYPE')))   
        except:
            raise TimeoutError('連線逾時，請關閉後重新操作')
        
        dropdown_by_value('_selOFFTYPE',day_or_hour)
        dropdown_by_value('_selLEAVETYPE',kind)
        dropdown_by_value('_selTASKDATE',date_num)

        #略過時段起迄
        if time_st == 'None':
            #新增
            click_by_id('_btnInsert')
            alert_click()   
            continue

        #時段(迄)先結束後開始
        dropdown_by_index('_selDATEE',date_num_en)
        dropdown_by_value('_selTASKHOURE',time_en)

        #時段(起)
        dropdown_by_index('_selDATES',date_num_st)
        dropdown_by_value('_selTASKHOURS',time_st)

        #新增
        click_by_id('_btnInsert')
        alert_click()

#刪除重複登打
def duplication_clear():
    
    for m in range(200,3,-1):
        matrix = []
        matrixx = []
        try:
            for n in range(1,6):
                xp = f'//*[@id="frm"]/table/tbody/tr[4]/td/table[1]/tbody/tr[{m}]/td[{n}]'
                temp = driver.find_element(By.XPATH, xp).text
                matrix.append(temp)

                xp2 = f'//*[@id="frm"]/table/tbody/tr[4]/td/table[1]/tbody/tr[{m-1}]/td[{n}]'
                temp = driver.find_element(By.XPATH, xp2).text
                matrixx.append(temp)
            
            if matrix == matrixx:
                print(f'刪除 {m}：{matrix}')
                xpp = f'/html/body/form/table/tbody/tr[4]/td/table[1]/tbody/tr[{m}]/td[10]/input[2]'
                del_button = driver.find_element(By.XPATH,xpp)
                del_button.click()
                time.sleep(0.2)
                alert_click()
        except:
            continue
    return 0

#刪除全部
def del_all():
    for m in range(200,3,-1):
        matrix = []
        try:
            for n in range(1,6):
                xp = f'//*[@id="frm"]/table/tbody/tr[4]/td/table[1]/tbody/tr[{m}]/td[{n}]'
                temp = driver.find_element(By.XPATH, xp).text
                matrix.append(temp)

            print(f'刪除{m-3}：{matrix}')
            xpp = f'/html/body/form/table/tbody/tr[4]/td/table[1]/tbody/tr[{m}]/td[10]/input[2]'
            del_button = driver.find_element(By.XPATH,xpp)
            del_button.click()
            time.sleep(0.2)
            alert_click()
        except:
            continue
    return 0
def dropdown_by_value(id = str, value = str):
    dropdown1 = wait.until(EC.visibility_of_element_located((By.ID, id)))   
    dropdown_sheet1 = Select(dropdown1)
    dropdown_sheet1.select_by_value(value)

def dropdown_by_index(id = str, value = str):
    dropdown1 = wait.until(EC.visibility_of_element_located((By.ID, id)))   
    dropdown_sheet1 = Select(dropdown1)
    dropdown_sheet1.select_by_index(value)

def click_by_id(id = str):
    button1 = driver.find_element(By.ID, id)
    button1.click()

def click_by_name(name = str):
    button1 = driver.find_element(By.NAME, name)
    button1.click()

def str_line(show = str):
    max_len = 50
    dash_len = int((max_len - len(show))/2)
    dash = ''
    for i in range(dash_len):
        dash = dash + '-'
    show = dash + show + dash + '\n'
    
    return show
################################################主程式################################################
if __name__ == '__main__':

    print(str_line("嗨，我是打基準表小精靈5.1，來完成任務吧"))   #版本號更新於此
    time.sleep(0.5)

    #檢測指定Excel是否存在
    file_name = "個人基準表vvcd.xlsm"
    user_book = "日曆_輸入區"
    duty_book = "假別_設定區"
    #setting_name = "setting_py"
    exe_dir = get_exe_dir()
    file_path = os.path.join(exe_dir, file_name)

    check_msg = excel_check(file_path)
    if check_msg == '檔案可讀取':
        print(f'{file_name} 讀取成功\n')
    elif check_msg == '檔案不存在':
        print('找不到Excel，確認有沒有在這個資料夾吧')
        time.sleep(2)
        raise Exception
    elif check_msg == '檔案被占用':
        print('Excel開啟中，請先存檔關閉再執行喔')
        time.sleep(2)
        raise PermissionError

    #讀取Excel內容
    #data字典：[name,account,password,month,year,days]
    data = data_loading(file_path, user_book)
    print(f'輸入：{data['name']}-{data['year']}年{data['month']}月基準表\n')
    last_day = data['days']

    #讀取假別列表 
    duty_dict = setting_loading(file_path, duty_book)

    #選擇是否刪除
    if(data["isdel"] == "清起來"):
        print(f'你在【日曆_輸入區：H8】選擇了{data["isdel"]}，要刪除{data["year"]}年{data["month"]}月的個人基準表所有內容嗎?')
        isdel = int(input('輸入"1"開始刪除，或"0"繼續登打↵\n'))
    else:
        isdel = 0

    time.sleep(0.5)
    ################Chrome操作################
    print(str_line('基準表，啟動'))
    #開啟Chrome瀏覽器、勤務系統
    driver = webdriver.Chrome()
    driver.get('https://dutymgt.tyfd.gov.tw/tyfd119/login119')

    #登入操作
    username = driver.find_element(By.ID,"_txtUsername")
    password = driver.find_element(By.ID,"_txtPassword")
    username.send_keys(data['account'])
    password.send_keys(data['password'])

    click_by_name('login')  #點選登入
    try:
        time.sleep(0.5)
        frameM = driver.find_element(By.NAME, 'ehrFrame')
    except:
        raise Exception('帳密錯誤，請輸入後重啟')

    print('登入成功')
    wait = WebDriverWait(driver, 10)  # 最長等待 10 秒
    
    #切換到選單Frame|#frameset是組合，不是Frame
    frameM = driver.find_element(By.NAME, 'ehrFrame')
    driver.switch_to.frame(frameM)
    frameL1 = driver.find_element(By.NAME, 'sidemenuFrame')
    driver.switch_to.frame(frameL1)
    frameL2 = driver.find_element(By.NAME, 'contentSidemenu')
    driver.switch_to.frame(frameL2)

    click_by_name('nodeIcon1')   #轉換左方選單
    click_by_id('itemIcon2')   #勤務基準表按鈕

    #轉換右方主要內容
    driver.switch_to.parent_frame()
    driver.switch_to.parent_frame()
    frameR1 = driver.find_element(By.NAME, 'contentFrame')
    driver.switch_to.frame(frameR1)
    frameR2 = driver.find_element(By.NAME, 'mainFrame')
    driver.switch_to.frame(frameR2)
    
    #點選時數上限警告
    alert_click()
    #點選時數上限警告
    alert_click()
    
    #查詢月份
    try:
        sheet_year = wait.until(EC.visibility_of_element_located((By.ID, '_selYEAR')))
    except:
        raise TimeoutError('連線逾時，請關閉後重新操作')

    dropdown_by_value('_selYEAR',data['year'])
    dropdown_by_value('_selMONTH',data['month'])
    click_by_id('_btnQuery')    #點選查詢
    alert_click()   #點選時數上限警告
    sheet_year = wait.until(EC.visibility_of_element_located((By.ID, '_selYEAR')))  #等待

    #讀取Excel個人資料、操作網頁、點選個人連結
    my_name = data['name']
    my_name_NAME = ''
    for n in range(1,70):
        ff_name = driver.find_element(By.NAME, '_td_' + str(n) + '_0')
        if ff_name.text != my_name:
            continue
        else:
            print(f"你是單位第{n}人\n")
            my_name_NAME = '_td_' + str(n) + '_0'
            break
    click_by_name(my_name_NAME)   #點進個人視窗

    #切換到彈出小視窗
    window_handles = driver.window_handles
    popup_window_handle = window_handles[-1]
    driver.switch_to.window(popup_window_handle)

    #全刪或登打
    if(isdel):
        del_all()
        print(str_line('刪除完成，記得確認喔'))
    else:  
        duty_loading(file_path, user_book, last_day)
        print('\n重複內容檢查中...')   #刪除重複內容
        duplication_clear()
        print(str_line('登打完成，記得確認喔'))
    time.sleep(0.5)
    pau = input('輸入任意鍵結束')
    driver.close()